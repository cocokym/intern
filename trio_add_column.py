from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange

def add_columns_to_trio(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # Insert 'Reportable Variant' at column A
    ws.insert_cols(1)
    ws.cell(row=2, column=1, value='Reportable Variant')

    # Adjust all merged cells
    for merged_cell in list(ws.merged_cells.ranges):
        ws.merged_cells.remove(merged_cell)
        new_range = CellRange(min_col=merged_cell.min_col + 1, min_row=merged_cell.min_row,
                              max_col=merged_cell.max_col + 1, max_row=merged_cell.max_row)
        ws.merged_cells.add(new_range)

    # Insert new columns between 'flags' and 'Proband (IM673)'
    new_columns = [
        'IGV review (True / False call)', 'Zyogosity', 'Phenotype',
        'First review and comment', 'Second review and comment on reportable variant',
        'Special remarks'
    ]
    for i, col in enumerate(new_columns, start=6):
        ws.insert_cols(i)
        ws.cell(row=2, column=i, value=col)

    # Save the modified workbook
    wb.save(file_path)